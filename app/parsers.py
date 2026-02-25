from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


def parse_txt_or_md(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")


def parse_pdf(path: Path) -> str:
    reader = PdfReader(str(path))
    pages = [page.extract_text() or "" for page in reader.pages]
    return "\n".join(pages).strip()


def parse_xlsx(path: Path) -> str:
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    lines: list[str] = []
    for sheet in wb.worksheets:
        lines.append(f"[Arkusz: {sheet.title}]")
        for row in sheet.iter_rows(values_only=True):
            row_values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if row_values:
                lines.append(" | ".join(row_values))
    wb.close()
    return "\n".join(lines).strip()


def extract_text_for_file(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".txt", ".md"}:
        return parse_txt_or_md(path)
    if suffix == ".pdf":
        return parse_pdf(path)
    if suffix == ".xlsx":
        return parse_xlsx(path)
    raise ValueError(f"Unsupported extension: {suffix}")
